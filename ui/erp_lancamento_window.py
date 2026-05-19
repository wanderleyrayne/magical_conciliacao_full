import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import warnings
import pandas as pd

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning,
    module="openpyxl"
)
import numpy as np
import json
from datetime import datetime, date as _date
from pathlib import Path

from utils.paths import app_path, user_data_path
from database.repository import SystemRepository


TIPO_DESPESA_MAP = {
    "CMV":                  2,
    "DespesasPrestadores":  2,
    "DespesasGerais":       2,
    "DespesasOcupacional":  2,
    "OutrasDespesas":       2,
}

FORMA_PGTO_MAP = {
    "PIX":      3,
    "TED":      2,
    "BOLETO":   4,
    "DINHEIRO": 1,
}
FORMA_PGTO_DEFAULT = 1

CATEGORIA_MAP = {
    "Brigadista":                    "16599",
    "Cerimonialista":                "16799",
    "Chef":                          "15699",
    "Decorador":                     "15599",
    "Eletricista":                   "16699",
    "Gerente de Evento":             "15999",
    "Mão de Obra Especializada":     "16999",
    "Recepcionista":                 "15799",
    "Segurança e Manobrista":        "16899",
    "Manobrista":                    "16899",
    "Bar":                           "17699",
    "Bebidas":                       "17099",
    "Bolo":                          "17199",
    "Buffet":                        "17299",
    "Doce":                          "17399",
    "Produtos Alimentícios - Outros":"17499",
    "Degustação":                    "17599",
    "Decoração":                     "17899",
    "Limpeza":                       "18299",
    "Som & Imagem":                  "18099",
    "Iluminação":                    "18199",
    "Combustível":                   "14099",
    "Frete":                         "17799",
    "Translado":                     "15299",
    "Transporte Cargas":             "2082717",
    "Aluguel Mobiliário":            "17999",
    "Prestação de Serviços":         "2082704",
    "Prestação de Serviços - T":     "15499",
    "Prestação de Serviços - A":     "15399",
    "Prestação de Serviços - CB":    "21299",
    "Repasse":                       "14199",
    "Repasse - P":                   "2082708",
    "Repasse - SG":                  "21199",
    "Salário":                       "2082716",
    "Remuneração - Vendedores":      "20899",
    "Remuneração - Planejadoras":    "20999",
    "Remuneração - SDR":             "21099",
    "Férias":                        "21499",
    "Rescisões Trabalhistas":        "15099",
    "Pró-Labore Sócios":             "20699",
    "Assistência Contábil & BPO":    "14299",
    "Assistência Jurídica":          "19599",
    "Cartório":                      "13999",
    "Copa & Cozinha":                "19699",
    "Estornos - Lançamentos Indevidos": "2082706",
    "Instalações Comerciais":        "19999",
    "Locação de Equipamentos":       "19799",
    "Manutenção & Conservação":      "14799",
    "Material de Escritório":        "20099",
    "Materiais de Limpeza & Higiene":"14699",
    "Móveis & Utensílios":           "19899",
    "Obras & Reformas":              "19499",
    "Passagem aéreas":               "14899",
    "Serviços Gerais":               "20599",
    "Tecnologia da Informação":      "20199",
    "Telefonia & Internet":          "20299",
    "Elaboração Conteúdo":           "14999",
    "Estrutura de Marketing":        "16299",
    "Marketing & Publicidade":       "2082699",
    "Tráfego Pago":                  "2081799",
    "Água e esgoto":                 "13599",
    "Alarmes":                       "19399",
    "Aluguel":                       "13699",
    "Aquisição de Equipamentos":     "13799",
    "Condomínio":                    "18699",
    "Energia Elétrica":              "18999",
    "Gás":                           "19099",
    "Gerador":                       "19199",
    "IPTU":                          "2080799",
    "Seguros":                       "19299",
    "Taxas":                         "18899",
    "COFINS":                        "2082720",
    "CSLL":                          "2080699",
    "CSLL & IRPJ":                   "2082718",
    "FGTS":                          "2080299",
    "ICMS":                          "2080399",
    "IMPOSTO":                       "2081399",
    "INSS":                          "2080499",
    "IOF":                           "15899",
    "IRPJ":                          "2082723",
    "ISS":                           "2081199",
    "PIS":                           "2082724",
    "PIS & COFINS":                  "2082719",
    "SIMPLES NACIONAL":              "2082721",
    "Tarifa Boleto":                 "18499",
    "Tarifa Cartão Crédito":         "2082499",
    "Tarifa Cartão Débito":          "2082399",
    "Tarifa PIX":                    "18599",
    "Tarifas Bancárias":             "15199",
    "Antecipação - Degustação":      "2082712",
    "Antecipações & Retiradas Sócios":"20499",
    "Arrendamento":                  "13899",
    "Comissão Magical":              "2081499",
    "Despesas com Sócios":           "20399",
    "Devoluções de Vendas":          "13499",
    "Empréstimos":                   "14399",
    "Investimentos":                 "14499",
    "Juros":                         "14599",
    "Participação Sócios":           "2082599",
    "Receita - Sinal":               "13099",
    "Receita - Recorrentes":         "13199",
    "Receita - Integralização":      "13399",
    "Receita - Aplicações":          "13299",
    "Receita - Financeiras":         "16099",
    "Receita - Quitação":            "16099",
    "Receita - Serviços":            "2081599",
    "Receita - Opcional Pós Venda":  "2082709",
    "Estrutura de Marketing (R)":    "2082299",
    "Crédito Indevido":              "2082707",
    "DespFixaVariável":              "2082704",
}
CATEGORIA_DEFAULT = "2082704"


class ErpLancamentoWindow:
    COL_DATA          = "DATA"
    COL_TIPO          = "TIPO DESPESA\nMeEventos"
    COL_ID_EVENTO     = "ID\nMeEventos"
    COL_CATEGORIA     = "CATEGORIA\nMeEventos"
    COL_CATEGORIA_ALT = "CATEGORIA MeEventos"
    COL_DETALHE       = "DETALHE PAGAMENTO"
    COL_VALOR         = "VALOR"
    COL_FORNECEDOR    = "FORNECEDOR"
    COL_FAVORECIDO    = "FAVORECIDO"
    COL_FORMA_PGTO    = "FORMA\nPGTO"
    COL_LANCAMENTO    = "PARA LANÇAMENTO DO FINANCEIRO NO Me Eventos"

    STATUS_OK          = "✓ Pronto"
    STATUS_ATENCAO     = "⚠ Atenção"
    STATUS_BLOQUEIO    = "✗ Bloqueado"
    STATUS_JA_LANCADO  = "↩ Já lançado"

    def __init__(self, master):
        self.top = tk.Toplevel(master)
        self.top.title("Lançamento no ERP — MeEventos")
        self.top.geometry("1300x780")
        self.top.minsize(1100, 600)

        try:
            icon_path = app_path("assets", "icon.ico")
            if icon_path.exists():
                self.top.iconbitmap(str(icon_path))
        except Exception:
            pass

        self.repo = SystemRepository(str(user_data_path("data", "conciliacao.db")))

        self.df_raw      = None
        self.df_preview  = None
        self.file_path   = None
        self._categoria_map_local    = {}
        self._payment_methods_local  = {}  # forma_pgto → id da API
        self._futuras_info   = None
        self._data_future    = None
        self._avisos_carregamento = []

        self._build_layout()
        self._load_api_settings()
        self.top.protocol("WM_DELETE_WINDOW", self._on_close)

    # =========================================================================
    # LAYOUT
    # =========================================================================

    def _build_layout(self):
        header = tk.Frame(self.top, bg="#1e293b", height=48)
        header.pack(fill="x")
        tk.Label(
            header, text="Lançamento de Despesas — ERP MeEventos",
            fg="white", bg="#1e293b", font=("Arial", 12, "bold")
        ).pack(side="left", padx=12, pady=10)

        api_frame = tk.LabelFrame(self.top, text="Parceiro e modo", padx=10, pady=8)
        api_frame.pack(fill="x", padx=12, pady=(10, 0))

        tk.Label(api_frame, text="Parceiro:").grid(row=0, column=0, sticky="w", pady=4)

        from core.partner_rules import PARTNERS as _PARTNERS
        partner_names = [p["partner_name"] for p in _PARTNERS]

        self.partner_var = tk.StringVar()
        self.partner_combo = ttk.Combobox(
            api_frame, textvariable=self.partner_var,
            values=partner_names, state="readonly", width=28
        )
        self.partner_combo.grid(row=0, column=1, sticky="w", padx=8, pady=4)
        self.partner_combo.bind("<<ComboboxSelected>>", self._on_partner_selected)

        self.api_status_label = tk.Label(
            api_frame, text="Selecione o parceiro para carregar a configuração da API.",
            fg="#475569", anchor="w"
        )
        self.api_status_label.grid(row=0, column=2, columnspan=2, sticky="w", padx=(12, 0))

        tk.Label(api_frame, text="Modo:").grid(row=1, column=0, sticky="w", pady=4)
        self.dry_run_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            api_frame,
            text="Simulação (não envia para a API — só valida)",
            variable=self.dry_run_var
        ).grid(row=1, column=1, columnspan=3, sticky="w", padx=8)

        self._current_url   = ""
        self._current_token = ""

        file_frame = tk.LabelFrame(self.top, text="Planilha de despesas", padx=10, pady=8)
        file_frame.pack(fill="x", padx=12, pady=(8, 0))

        self.file_label = tk.Label(file_frame, text="Nenhuma planilha selecionada", fg="#64748b", anchor="w")
        self.file_label.pack(side="left", fill="x", expand=True)

        ttk.Button(file_frame, text="Selecionar planilha", command=self._select_file).pack(side="left", padx=6)
        ttk.Button(file_frame, text="↺ Revalidar", command=self._revalidate).pack(side="left")

        filter_frame = tk.Frame(self.top)
        filter_frame.pack(fill="x", padx=12, pady=(8, 0))

        tk.Label(filter_frame, text="Exibir:").pack(side="left")
        self.filter_status_var = tk.StringVar(value="TODOS")
        ttk.Combobox(
            filter_frame,
            textvariable=self.filter_status_var,
            values=["TODOS", self.STATUS_OK, self.STATUS_ATENCAO,
                    self.STATUS_BLOQUEIO, self.STATUS_JA_LANCADO],
            state="readonly", width=22
        ).pack(side="left", padx=6)
        self.filter_status_var.trace_add("write", lambda *a: self._refresh_preview())

        self.summary_label = tk.Label(filter_frame, text="", font=("Arial", 9, "bold"))
        self.summary_label.pack(side="left", padx=12)

        table_frame = tk.Frame(self.top)
        table_frame.pack(fill="both", expand=True, padx=12, pady=(6, 0))

        cols = ("status", "linha", "data", "id_evento", "categoria",
                "descricao", "valor", "forma_pgto", "aviso")

        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="extended")
        self.tree.pack(side="left", fill="both", expand=True)

        sy = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        sy.pack(side="right", fill="y")
        sx = ttk.Scrollbar(self.top, orient="horizontal", command=self.tree.xview)
        sx.pack(fill="x", padx=12)
        self.tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)

        headings = {
            "status":    "Status",
            "linha":     "Linha",
            "data":      "Data",
            "id_evento": "ID Evento",
            "categoria": "Categoria [ID]",
            "descricao": "Descrição (enviada à API)",
            "valor":     "Valor",
            "forma_pgto":"Modo Pgto",
            "aviso":     "Aviso / Observação",
        }
        widths = {
            "status": 100, "linha": 55, "data": 95,
            "id_evento": 80, "categoria": 180, "descricao": 280,
            "valor": 100, "forma_pgto": 80, "aviso": 300,
        }
        for col in cols:
            self.tree.heading(col, text=headings[col])
            self.tree.column(col, width=widths[col], anchor="w")

        self.tree.tag_configure(self.STATUS_OK,         background="#dff3e3")
        self.tree.tag_configure(self.STATUS_ATENCAO,    background="#fff3cd")
        self.tree.tag_configure(self.STATUS_BLOQUEIO,   background="#f8d7da")
        self.tree.tag_configure(self.STATUS_JA_LANCADO, background="#e0e7ff", foreground="#3730a3")
        self.tree.tag_configure("LANCADO",              background="#dbeafe")
        self.tree.tag_configure("ERRO_API",             background="#f8d7da")

        self.tree.bind("<Double-1>", self._on_row_double_click)

        footer = tk.Frame(self.top, bd=1, relief="solid", padx=10, pady=8)
        footer.pack(fill="x", padx=12, pady=(6, 12))

        self.btn_lancar = ttk.Button(
            footer, text="Lançar no ERP →", command=self._confirm_and_launch,
            state="disabled"
        )
        self.btn_lancar.pack(side="right", padx=6)

        ttk.Button(footer, text="Fechar", command=self.top.destroy).pack(side="right")

        ttk.Separator(footer, orient="vertical").pack(side="right", fill="y", padx=8)

        def _open_history():
            try:
                from ui.erp_history_window import ErpHistoryWindow
                ErpHistoryWindow(self.top, self.repo)
            except Exception as exc:
                messagebox.showerror("Erro", str(exc), parent=self.top)

        ttk.Button(footer, text="📋 Histórico", command=_open_history).pack(side="right")

        self.progress = ttk.Progressbar(footer, orient="horizontal", length=200, mode="determinate")
        self.progress.pack(side="left", padx=(0, 10))

        self.status_label = tk.Label(footer, text="Selecione uma planilha para começar.", fg="#475569")
        self.status_label.pack(side="left")

        self.total_valor_label = tk.Label(
            footer, text="", font=("Arial", 10, "bold"), fg="#1e293b")
        self.total_valor_label.pack(side="right", padx=(10, 0))

    # =========================================================================
    # HELPERS DE UI
    # =========================================================================

    def _on_partner_selected(self, event=None):
        name = self.partner_var.get()
        if not name:
            return
        safe = name.lower().replace(" ", "_")
        url   = self.repo.get_setting(f"erp_api_url_{safe}")   or ""
        token = self.repo.get_setting(f"erp_api_token_{safe}") or ""
        self._current_url   = url
        self._current_token = token

        # Limpa cache de eventos e categorias ao trocar de parceiro
        # (cada casa tem eventos e IDs diferentes)
        self._event_date_cache       = {}
        self._categoria_map_local    = {}
        self._payment_methods_local  = {}
        self.df_raw     = None
        self.df_preview = None
        self.file_path  = None
        self.file_label.config(text="Nenhuma planilha selecionada", fg="#64748b")
        self.btn_lancar.config(state="disabled")
        self.summary_label.config(text="")
        self.total_valor_label.config(text="")
        self._set_status("Parceiro alterado — selecione uma planilha.")
        for item in self.tree.get_children():
            self.tree.delete(item)

        if url and token:
            self.api_status_label.config(
                text=f"API configurada ✓  ({url})", fg="#166534")
            # Carrega cache de formas de pagamento
            try:
                import json as _j
                cached_pm = self.repo.get_setting(f"erp_payment_methods_{safe}")
                if cached_pm:
                    self._payment_methods_local = _j.loads(cached_pm)
            except Exception:
                pass
            import threading as _t
            _t.Thread(target=self._load_categories,
                      args=(url, token, safe), daemon=True).start()
        elif url:
            self.api_status_label.config(
                text="URL carregada, mas token não configurado.", fg="#92400e")
        else:
            self.api_status_label.config(
                text="Parceiro sem configuração de API. Acesse Configurações → API MeEventos.",
                fg="#b91c1c")

    def _load_categories(self, url_base: str, token: str, safe_name: str):
        import json as _json
        from datetime import datetime, timedelta

        cache_key    = f"erp_categories_{safe_name}"
        cache_ts_key = f"erp_categories_ts_{safe_name}"

        try:
            cached    = self.repo.get_setting(cache_key)
            cached_ts = self.repo.get_setting(cache_ts_key)
            if cached and cached_ts:
                ts = datetime.fromisoformat(cached_ts)
                if datetime.now() - ts < timedelta(days=7):
                    self._categoria_map_local = _json.loads(cached)
                    return
        except Exception:
            pass

        try:
            import requests as _req
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type":  "application/json",
                "Accept":        "application/json",
            }

            # Busca categorias
            categoria_map = {}
            for tipo in ("despesas", "receitas"):
                page = 1
                while True:
                    resp = _req.get(
                        f"{url_base.rstrip('/')}/api/v1/financial-categories",
                        params={"tipo": tipo, "page": page, "page_size": 100},
                        headers=headers, timeout=10
                    )
                    if resp.status_code != 200:
                        break
                    data = resp.json()
                    for cat in data.get("data", []):
                        nome = str(cat.get("nome") or "").strip()
                        cid  = str(cat.get("id")   or "").strip()
                        if nome and cid:
                            categoria_map[nome] = cid
                    pagination = data.get("pagination", {})
                    if page >= pagination.get("total_page", 1):
                        break
                    page += 1

            if categoria_map:
                self._categoria_map_local = categoria_map
                self.repo.save_setting(cache_key, _json.dumps(categoria_map))
                self.repo.save_setting(cache_ts_key, datetime.now().isoformat())

            # Busca formas de pagamento
            pm_cache_key    = f"erp_payment_methods_{safe_name}"
            pm_cache_ts_key = f"erp_payment_methods_ts_{safe_name}"
            resp_pm = _req.get(
                f"{url_base.rstrip('/')}/api/v1/payment-methods",
                headers=headers, timeout=10
            )
            if resp_pm.status_code == 200:
                pm_map = {}
                for pm in resp_pm.json().get("data", []):
                    nome = str(pm.get("nome") or "").strip()
                    pid  = str(pm.get("id")   or "").strip()
                    if nome and pid:
                        # Mapeia nome exato e variacoes comuns
                        pm_map[nome.upper()] = pid
                        # Aliases comuns
                        if "PIX" in nome.upper():
                            pm_map["PIX"] = pid
                        elif "TRANSFER" in nome.upper() and "INTERN" not in nome.upper():
                            pm_map["TED"] = pid
                            pm_map["DOC"] = pid
                            pm_map["TRANSFERENCIA"] = pid
                        elif "BOLETO" in nome.upper():
                            pm_map["BOLETO"] = pid
                        elif "DINHEIRO" in nome.upper():
                            pm_map["DINHEIRO"] = pid
                        elif "DEBITO AUTO" in nome.upper():
                            pm_map["DEBITO"] = pid
                        elif "DEPOSITO" in nome.upper():
                            pm_map["DEPOSITO"] = pid

                if pm_map:
                    self._payment_methods_local = pm_map
                    self.repo.save_setting(pm_cache_key, _json.dumps(pm_map))
                    self.repo.save_setting(pm_cache_ts_key, datetime.now().isoformat())

        except Exception:
            pass

    def _load_api_settings(self):
        try:
            last = self.repo.get_setting("erp_last_partner")
            if last and last in [p["partner_name"] for p in __import__(
                    "core.partner_rules", fromlist=["PARTNERS"]).PARTNERS]:
                self.partner_var.set(last)
                self._on_partner_selected()
        except Exception:
            pass

    def _save_api_settings(self):
        try:
            self.repo.save_setting("erp_last_partner", self.partner_var.get())
        except Exception:
            pass

    def _parse_date_str(self, text: str) -> str | None:
        import re as _re
        text = text.strip()

        m = _re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', text)
        if m:
            return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"

        m = _re.match(r'^(\d{4})-(\d{2})-(\d{2})$', text)
        if m:
            return text

        m = _re.match(r'^(\d{1,2})-(\d{1,2})-(\d{4})$', text)
        if m:
            return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"

        m = _re.match(r'^(\d{2})(\d{2})/(\d{4})$', text)
        if m:
            return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"

        m = _re.match(r'^(\d{2})(\d{2})/(\d{2})$', text)
        if m:
            return f"20{m.group(3)}-{m.group(2)}-{m.group(1)}"

        try:
            dt = pd.to_datetime(text, dayfirst=True, errors="coerce")
            if not pd.isna(dt):
                return dt.strftime("%Y-%m-%d")
        except Exception:
            pass

        return None

    def _resolve_event_by_date(self, data_fmt: str, avisos: list) -> tuple:
        if not hasattr(self, "_event_date_cache"):
            self._event_date_cache = {}

        if data_fmt in self._event_date_cache:
            cached = self._event_date_cache[data_fmt]
            if cached:
                return cached, f"{cached} ({data_fmt})"
            return None, ""

        url_base = getattr(self, "_current_url", "").rstrip("/")
        token    = getattr(self, "_current_token", "")

        if not url_base or not token:
            self._event_date_cache[data_fmt] = None
            return None, ""

        try:
            import requests as _req
            resp = _req.get(
                f"{url_base}/api/v1/events",
                params={"start": data_fmt, "end": data_fmt},
                headers={
                    "Authorization": f"Bearer {token}",
                    "Accept": "application/json",
                },
                timeout=8,
            )
            if resp.status_code == 200:
                data = resp.json().get("data", [])
                if data:
                    evento_id = str(data[0].get("id", ""))
                    nome      = str(data[0].get("nomeevento", "")).strip()[:40]
                    if len(data) > 1:
                        avisos.append(
                            f"{len(data)} eventos em {data_fmt} — vinculado ao primeiro: {nome}")
                    self._event_date_cache[data_fmt] = evento_id
                    return evento_id, f"{evento_id} ({nome})"
                else:
                    self._event_date_cache[data_fmt] = None
                    return None, ""
            else:
                self._event_date_cache[data_fmt] = None
                return None, ""
        except Exception:
            pass

        self._event_date_cache[data_fmt] = None
        return None, ""

    def _revalidate(self):
        self._event_date_cache = {}
        self._load_and_validate()

    def _on_close(self):
        self._save_api_settings()
        self.top.destroy()

    def _set_status(self, text, color="#475569"):
        self.status_label.config(text=text, fg=color)

    def _on_row_double_click(self, event):
        item = self.tree.identify_row(event.y)
        if not item:
            return

        vals = self.tree.item(item, "values")
        if not vals:
            return

        try:
            linha_excel = int(vals[1])
            idx_original = linha_excel - 5
        except (ValueError, IndexError):
            return

        if self.df_preview is None or self.df_preview.empty:
            return

        matches = self.df_preview[self.df_preview["_idx"] == idx_original]
        if matches.empty:
            return

        try:
            row = matches.iloc[0]
        except (IndexError, KeyError):
            return
        payload = row.get("_payload", {})
        status  = row.get("_status", "")
        aviso   = row.get("_aviso", "")

        win = tk.Toplevel(self.top)
        win.title(f"Payload — Linha {linha_excel}")
        win.geometry("560x460")
        win.resizable(True, True)
        win.grab_set()

        bg_colors = {
            self.STATUS_OK:       "#166534",
            self.STATUS_ATENCAO:  "#92400e",
            self.STATUS_BLOQUEIO: "#991b1b",
            "✓ Lançado":          "#1e40af",
            "✗ Erro API":         "#991b1b",
        }
        bg = bg_colors.get(status, "#1e293b")
        hdr = tk.Frame(win, bg=bg)
        hdr.pack(fill="x")
        tk.Label(hdr, text=f"Linha {linha_excel} — {status}",
                 fg="white", bg=bg, font=("Arial", 10, "bold")).pack(
            side="left", padx=12, pady=8)

        if aviso:
            tk.Label(win, text=f"⚠  {aviso}", fg="#92400e", anchor="w",
                     font=("Arial", 9), wraplength=520).pack(
                fill="x", padx=12, pady=(8, 0))

        import json as _json
        payload_txt = _json.dumps(payload, ensure_ascii=False, indent=2)

        frame_txt = tk.Frame(win)
        frame_txt.pack(fill="both", expand=True, padx=12, pady=8)

        txt = tk.Text(frame_txt, wrap="word", font=("Courier New", 10),
                      bg="#f8fafc", relief="flat", bd=0)
        txt.pack(side="left", fill="both", expand=True)

        sb = ttk.Scrollbar(frame_txt, orient="vertical", command=txt.yview)
        sb.pack(side="right", fill="y")
        txt.configure(yscrollcommand=sb.set)

        txt.insert("1.0", payload_txt)
        txt.config(state="disabled")

        btn_frame = tk.Frame(win, padx=12, pady=8)
        btn_frame.pack(fill="x")

        def copiar():
            win.clipboard_clear()
            win.clipboard_append(txt.get("1.0", "end-1c"))
            btn_copy.config(text="\u2713 Copiado!")
            win.after(1500, lambda: btn_copy.config(text="Copiar JSON"))

        def editar():
            txt.config(state="normal", bg="white",
                       highlightthickness=1, highlightbackground="#f59e0b")
            modo_lbl.config(text="Modo: EDICAO — edite o JSON e clique em Salvar",
                            fg="#b45309", font=("Arial", 8, "bold"))
            btn_editar.config(state="disabled")
            btn_salvar.config(state="normal")
            txt.focus_set()

        def salvar():
            import json as _j2
            novo_txt = txt.get("1.0", "end-1c").strip()
            try:
                novo_payload = _j2.loads(novo_txt)
            except Exception as e:
                messagebox.showerror("JSON invalido",
                    f"Erro: {e}", parent=win)
                return

            df_idx = matches.index[0]
            self.df_preview.at[df_idx, "_payload"] = novo_payload

            if "idcategoria" in novo_payload:
                self.df_preview.at[df_idx, "_id_categoria"] = novo_payload["idcategoria"]
            if "valor" in novo_payload:
                self.df_preview.at[df_idx, "_valor"] = float(novo_payload["valor"])
            if "datapagamento" in novo_payload:
                self.df_preview.at[df_idx, "_data_fmt"] = novo_payload["datapagamento"]

            aviso_atual = str(self.df_preview.at[df_idx, "_aviso"])
            self.df_preview.at[df_idx, "_aviso"] = "Editado manualmente"
            if self.df_preview.at[df_idx, "_status"] in (
                    self.STATUS_ATENCAO, self.STATUS_JA_LANCADO):
                self.df_preview.at[df_idx, "_status"] = self.STATUS_OK

            self._refresh_preview()
            if hasattr(self, "_atualizar_summary"):
                self._atualizar_summary()

            txt.config(state="disabled", bg="#f8fafc", highlightthickness=0)
            modo_lbl.config(text="\u2705 Salvo! Payload atualizado.",
                            fg="#166534", font=("Arial", 8, "bold"))
            btn_salvar.config(state="disabled")
            btn_editar.config(state="normal")

        # Label modo (inserido antes dos botoes)
        modo_lbl = tk.Label(win, text="Modo: visualizacao  |  Clique em Editar para modificar",
                            fg="#64748b", font=("Arial", 8), anchor="w")
        modo_lbl.pack(fill="x", padx=12, pady=(0,4))

        btn_copy   = ttk.Button(btn_frame, text="Copiar JSON", command=copiar)
        btn_copy.pack(side="left")
        btn_editar = ttk.Button(btn_frame, text="\u270f\ufe0f Editar", command=editar)
        btn_editar.pack(side="left", padx=6)
        btn_salvar = ttk.Button(btn_frame, text="\U0001f4be Salvar",
                                command=salvar, state="disabled")
        btn_salvar.pack(side="left")
        ttk.Button(btn_frame, text="Fechar", command=win.destroy).pack(side="right")


    # =========================================================================
    # SELEÇÃO E LEITURA DA PLANILHA
    # =========================================================================

    def _select_file(self):
        path = filedialog.askopenfilename(
            title="Selecionar planilha de despesas",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
        )
        if path:
            self._load_file_from_path(path)

    def _load_file_from_path(self, path: str):
        self.file_path = path
        self.file_label.config(text=Path(path).name, fg="#166534")
        self._event_date_cache = {}
        self._load_and_validate()

    def _load_and_validate(self):
        if not self.file_path:
            messagebox.showwarning("Aviso", "Selecione uma planilha primeiro.")
            return

        self._set_status("Lendo planilha...")
        self.btn_lancar.config(state="disabled")

        threading.Thread(target=self._load_worker, daemon=True).start()

    def _find_header_row(self, df: pd.DataFrame) -> int:
        for i in range(min(10, len(df))):
            row_vals = [str(v).strip() for v in df.iloc[i].tolist()]
            if "DATA" in row_vals:
                return i
        raise ValueError(
            "Cabeçalho 'DATA' não encontrado nas primeiras 10 linhas da aba 'Despesas'.\n"
            "Verifique se o arquivo selecionado é a planilha de despesas correta."
        )

    def _build_clean_headers(self, raw_headers: list) -> list:
        clean = []
        seen = {}
        for i, h in enumerate(raw_headers):
            key = str(h).strip() if h is not None and str(h).strip() not in ("nan", "") else ""
            if not key:
                key = f"_col_{i}"
            if key in seen:
                seen[key] += 1
                key = f"{key}.{seen[key]}"
            else:
                seen[key] = 0
            clean.append(key)
        return clean

    def _load_worker(self):
        try:
            # Tenta ler normalmente
            try:
                df_raw = pd.read_excel(self.file_path, sheet_name="Despesas", header=None)
            except Exception as e1:
                # Fallback 1: ignora stylesheet corrompido (Google Sheets / LibreOffice)
                if "stylesheet" in str(e1).lower() or "invalid xml" in str(e1).lower():
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(
                            self.file_path, data_only=True, keep_vba=False)
                        ws = (wb["Despesas"] if "Despesas" in wb.sheetnames
                              else wb.active)
                        data_rows = [[cell.value for cell in row] for row in ws.iter_rows()]
                        df_raw = pd.DataFrame(data_rows)
                    except Exception:
                        # Fallback 2: primeira aba disponivel
                        df_raw = pd.read_excel(
                            self.file_path, sheet_name=0, header=None,
                            engine="openpyxl")
                else:
                    raise e1

            header_row_idx = self._find_header_row(df_raw)
            headers = self._build_clean_headers(df_raw.iloc[header_row_idx].tolist())

            data = df_raw.iloc[header_row_idx + 1:].reset_index(drop=True)
            data.columns = headers

            if self.COL_CATEGORIA not in data.columns and self.COL_CATEGORIA_ALT in data.columns:
                data = data.rename(columns={self.COL_CATEGORIA_ALT: self.COL_CATEGORIA})

            keep = [
                self.COL_DATA, self.COL_TIPO, self.COL_ID_EVENTO,
                self.COL_CATEGORIA, self.COL_DETALHE, self.COL_VALOR,
                self.COL_FORNECEDOR, self.COL_FAVORECIDO, self.COL_FORMA_PGTO,
                self.COL_LANCAMENTO,
            ]
            missing = [c for c in keep if c not in data.columns]
            if missing:
                raise ValueError(
                    f"Coluna(s) não encontrada(s) na planilha: {missing}\n"
                    f"Colunas disponíveis: {[c for c in data.columns if not c.startswith('_col_')]}"
                )
            data = data[keep].copy()

            # Filtra linhas com valor válido
            data = data[data[self.COL_VALOR].notna()].copy()
            data = data[data[self.COL_VALOR].apply(
                lambda x: isinstance(x, (int, float)) and not np.isnan(float(x))
            )].copy()

            # Separa linhas sem data (avisa) e com data
            sem_data = data[data[self.COL_DATA].isna() |
                           data[self.COL_DATA].apply(
                               lambda x: str(x).strip() in ("", "nan", "None"))].copy()
            com_data = data[~data.index.isin(sem_data.index)].copy()

            # Filtra por data: hoje=OK, futuro=pergunta, passado=ignora
            hoje = _date.today()
            passadas = []
            futuras  = []
            validas  = []

            for idx, row in com_data.iterrows():
                try:
                    dt = pd.to_datetime(row[self.COL_DATA], dayfirst=True).date()
                    if dt < hoje:
                        passadas.append(idx)
                    elif dt > hoje:
                        futuras.append(idx)
                    else:
                        validas.append(idx)
                except Exception:
                    validas.append(idx)

            data_valida  = com_data.loc[validas].copy()
            data_future  = com_data.loc[futuras].copy()

            # Avisos
            avisos = []
            if sem_data is not None and not sem_data.empty:
                try:
                    vals_sem = sem_data[self.COL_VALOR].sum()
                    vals_fmt = (f"R$ {vals_sem:,.2f}"
                                .replace(",","X").replace(".",",").replace("X","."))
                except Exception:
                    vals_fmt = "?"
                avisos.append(
                    f"{len(sem_data)} linha(s) sem data encontrada(s) — {vals_fmt}.\n"
                    f"Verifique a planilha antes de lançar.")

            self._avisos_carregamento = avisos
            self._futuras_info = (len(futuras),
                sorted(set(str(com_data.loc[i, self.COL_DATA])[:10]
                           for i in futuras))) if futuras else None
            self._data_future  = data_future
            self._data_passada = com_data.loc[passadas].copy() if passadas else pd.DataFrame()

            # Junta validas + sem_data (sem_data ficará bloqueada no validate_row)
            data_final = pd.concat([data_valida, sem_data]).reset_index(drop=True)

            # Garante que formas de pagamento foram carregadas antes de validar
            if not self._payment_methods_local and self._current_url and self._current_token:
                try:
                    import requests as _req
                    safe = self.partner_var.get().lower().replace(" ","_") if hasattr(self, 'partner_var') else ""
                    pm_cached = self.repo.get_setting(f"erp_payment_methods_{safe}")
                    if pm_cached:
                        import json as _jj
                        self._payment_methods_local = _jj.loads(pm_cached)
                    else:
                        resp_pm = _req.get(
                            f"{self._current_url.rstrip('/')}/api/v1/payment-methods",
                            headers={"Authorization": f"Bearer {self._current_token}",
                                     "Content-Type": "application/json",
                                     "Accept": "application/json"},
                            timeout=10,
                        )
                        if resp_pm.status_code == 200:
                            pm_map = {}
                            for pm in resp_pm.json().get("data", []):
                                nome = str(pm.get("nome") or "").strip()
                                pid  = str(pm.get("id")   or "").strip()
                                if nome and pid:
                                    pm_map[nome.upper()] = pid
                                    if "PIX" in nome.upper():
                                        pm_map["PIX"] = pid
                                    elif "TRANSFER" in nome.upper() and "INTERN" not in nome.upper():
                                        pm_map["TED"] = pid
                                        pm_map["DOC"] = pid
                                        pm_map["TRANSFERENCIA"] = pid
                                    elif "BOLETO" in nome.upper():
                                        pm_map["BOLETO"] = pid
                                    elif "DINHEIRO" in nome.upper():
                                        pm_map["DINHEIRO"] = pid
                            self._payment_methods_local = pm_map
                            import json as _jj
                            if pm_map and safe:
                                self.repo.save_setting(f"erp_payment_methods_{safe}", _jj.dumps(pm_map))
                except Exception as e:
                    print(f"[ERP] Erro ao carregar payment-methods: {e}")

            print(f"[ERP] payment_methods_local: {self._payment_methods_local}")

            self.df_raw     = data_final
            self.df_preview = self._validate_all(data_final)

            self.top.after(0, self._on_load_done)

        except Exception as exc:
            self.top.after(0, lambda m=str(exc): self._set_status(
                f"Erro ao ler planilha: {m}", "#b91c1c"))

    def _on_load_done(self):
        # Avisos de carregamento
        for av in getattr(self, "_avisos_carregamento", []):
            messagebox.showwarning("Planilha", av, parent=self.top)

        # Pergunta sobre passadas — mostra lista com checkboxes
        data_passada = getattr(self, "_data_passada", None)
        if data_passada is not None and not data_passada.empty:
            self._popup_selecionar_passadas(data_passada)

        # Pergunta sobre futuras
        futuras_info = getattr(self, "_futuras_info", None)
        data_future  = getattr(self, "_data_future", None)
        if futuras_info and data_future is not None and not data_future.empty:
            qtd, datas = futuras_info
            datas_str  = ", ".join(datas[:5])
            incluir = messagebox.askyesno(
                "Datas Futuras",
                f"{qtd} linha(s) com data futura encontrada(s):\n{datas_str}\n\n"
                f"Deseja incluir essas linhas no lançamento?",
                parent=self.top
            )
            if incluir:
                extra = self._validate_all(data_future.reset_index(drop=True))
                self.df_preview = pd.concat(
                    [self.df_preview, extra], ignore_index=True)

        self._refresh_preview()
        total = len(self.df_preview)
        ok    = (self.df_preview["_status"] == self.STATUS_OK).sum()
        atenc = (self.df_preview["_status"] == self.STATUS_ATENCAO).sum()
        bloq  = (self.df_preview["_status"] == self.STATUS_BLOQUEIO).sum()
        jalc  = (self.df_preview["_status"] == self.STATUS_JA_LANCADO).sum()

        # Total apenas das linhas novas (nao ja lancadas)
        total_val = self.df_preview[
            self.df_preview["_status"].isin([self.STATUS_OK, self.STATUS_ATENCAO])
        ]["_valor"].sum()
        total_val_fmt = (f"R$ {total_val:,.2f}"
                         .replace(",","X").replace(".",",").replace("X","."))

        self.summary_label.config(
            text=f"Total: {total}  |  ✓ {ok}  ⚠ {atenc}  ✗ {bloq}  ↩ {jalc} já lançados"
        )
        self.total_valor_label.config(
            text=f"Total a lançar: {total_val_fmt}"
        )
        self._set_status(f"{total} despesas carregadas. Revise e clique em 'Lançar no ERP'.")
        if ok + atenc > 0:
            self.btn_lancar.config(state="normal")

    # =========================================================================
    # VALIDAÇÃO
    # =========================================================================

    def _validate_all(self, data: pd.DataFrame) -> pd.DataFrame:
        dup_indices = set()
        seen = {}

        for i, row in data.iterrows():
            try:
                data_val = pd.to_datetime(row.get(self.COL_DATA)).strftime("%Y-%m-%d")
            except Exception:
                data_val = str(row.get(self.COL_DATA) or "")

            descricao  = str(row.get(self.COL_LANCAMENTO) or "").strip()
            if not descricao or descricao in ("nan", "#REF!"):
                descricao = str(row.get(self.COL_DETALHE) or "").strip()

            fornecedor = str(row.get("FORNECEDOR") or row.get("FAVORECIDO") or "").strip()

            try:
                valor = round(abs(float(row.get(self.COL_VALOR) or 0)), 2)
            except Exception:
                valor = 0.0

            id_evento_raw = str(row.get(self.COL_ID_EVENTO) or "").strip()
            SEM_ID_VALORES = {"nan", "", "sem id", "preencher id", "preencher", "s/id", "n/a", "na", "-"}
            tem_id = id_evento_raw.lower() not in SEM_ID_VALORES
            id_evento_key = id_evento_raw if tem_id else "__sem_id__"

            favorecido = str(row.get("FAVORECIDO") or "").strip()

            if tem_id:
                chave = (id_evento_key, valor, fornecedor, favorecido, descricao)
            else:
                chave = (id_evento_key, valor, fornecedor, favorecido, descricao, str(i))

            if chave in seen:
                dup_indices.add(seen[chave])
                dup_indices.add(i)
            else:
                seen[chave] = i

        rows = []
        for i, row in data.iterrows():
            validated = self._validate_row(i, row, is_duplicate=(i in dup_indices))
            rows.append(validated)
        return pd.DataFrame(rows)

    def _validate_row(self, idx, row, is_duplicate=False) -> dict:
        avisos = []
        status = self.STATUS_OK

        if is_duplicate:
            avisos.append("Linha duplicada — mesma data, descrição e valor")
            status = self.STATUS_BLOQUEIO
        else:
            try:
                partner   = self.partner_var.get() if hasattr(self, 'partner_var') else ""
                file_name = Path(self.file_path).name if self.file_path else ""
                linha_num = idx + 5

                ja_lancado = self.repo.check_already_launched(
                    partner_name=partner,
                    file_name=file_name,
                    linha=linha_num,
                    descricao=str(row.get(self.COL_LANCAMENTO) or ""),
                    valor=abs(float(row.get(self.COL_VALOR) or 0)),
                )
                if ja_lancado:
                    avisos.append(
                        f"Já lançado em {str(ja_lancado.get('created_at',''))[:16]}"
                        f" | ID API: {ja_lancado.get('id_api','—')}"
                    )
                    status = self.STATUS_JA_LANCADO
            except Exception:
                pass

        # ── DATA ──────────────────────────────────────────────────────────────
        data_val = row.get(self.COL_DATA)
        data_str = str(data_val or "").strip()

        if data_str in ("", "nan", "None", "NaT"):
            data_fmt = ""
            avisos.append("Sem data — revise a planilha")
            status = self.STATUS_BLOQUEIO
        else:
            try:
                data_fmt = pd.to_datetime(data_val, dayfirst=True).strftime("%Y-%m-%d")
            except Exception:
                data_fmt = data_str
                avisos.append("Data inválida")
                status = self.STATUS_BLOQUEIO

        # ── TIPO DESPESA ───────────────────────────────────────────────────────
        tipo_raw = str(row.get(self.COL_TIPO, "") or "").strip()
        tipocobranca = TIPO_DESPESA_MAP.get(tipo_raw, 2)

        # ── ID EVENTO ─────────────────────────────────────────────────────────
        id_evento_raw = row.get(self.COL_ID_EVENTO)
        id_evento = None
        id_evento_display = ""

        if isinstance(id_evento_raw, (int, float)) and not np.isnan(float(id_evento_raw)):
            id_evento = int(id_evento_raw)
            id_evento_display = str(id_evento)
        elif isinstance(id_evento_raw, datetime):
            data_evento = id_evento_raw.strftime("%Y-%m-%d")
            id_evento, id_evento_display = self._resolve_event_by_date(data_evento, avisos)
            if not id_evento:
                id_evento_display = id_evento_raw.strftime("%d/%m/%Y")
                avisos.append(f"Data '{id_evento_display}' — nenhum evento encontrado")
                status = max(status, self.STATUS_ATENCAO, key=self._status_weight)
        elif str(id_evento_raw or "").strip().lower() in ("sem id", "preencher id", "preencher", "s/id", "nan", ""):
            id_evento_display = "Sem ID"
            avisos.append("Sem ID de evento — lançará sem vínculo")
            status = max(status, self.STATUS_ATENCAO, key=self._status_weight)
        else:
            raw_str = str(id_evento_raw).strip()
            data_fmt_ev = self._parse_date_str(raw_str)
            if data_fmt_ev:
                id_evento, id_evento_display = self._resolve_event_by_date(data_fmt_ev, avisos)
                if not id_evento:
                    id_evento_display = raw_str
                    avisos.append(f"Data '{raw_str}' — nenhum evento encontrado")
                    status = max(status, self.STATUS_ATENCAO, key=self._status_weight)
            else:
                try:
                    id_evento = int(float(raw_str))
                    id_evento_display = str(id_evento)
                except Exception:
                    id_evento_display = raw_str
                    avisos.append(f"ID pendente: '{raw_str}'")
                    status = max(status, self.STATUS_ATENCAO, key=self._status_weight)

        # ── VALOR ─────────────────────────────────────────────────────────────
        valor_raw = row.get(self.COL_VALOR, 0)
        valor = 0.0

        def _parse_valor(v):
            import re as _re
            if isinstance(v, (int, float)):
                try:
                    f = float(v)
                    return abs(f) if f != float('nan') else None
                except Exception:
                    return None
            s = str(v or "").strip()
            s = _re.sub(r'[R\$\s]', '', s)
            if not s or s in ("nan", "None", ""):
                return None
            if ',' in s:
                try:
                    return abs(float(s.replace(".", "").replace(",", ".")))
                except Exception:
                    pass
            try:
                return abs(float(s.replace(",", "")))
            except Exception:
                pass
            return None

        resultado = _parse_valor(valor_raw)
        if resultado and resultado > 0:
            valor = resultado
        else:
            avisos.append(f"Valor inválido ou zerado: '{valor_raw}'")
            status = self.STATUS_BLOQUEIO

        # ── DESCRIÇÃO ─────────────────────────────────────────────────────────
        lancamento = str(row.get(self.COL_LANCAMENTO, "") or "").strip()
        detalhe    = str(row.get(self.COL_DETALHE, "")    or "").strip()
        favorecido = str(row.get(self.COL_FAVORECIDO, "") or "").strip()

        if lancamento and lancamento not in ("nan",) and "#REF!" not in lancamento:
            descricao = lancamento.lstrip(" |").strip()
        else:
            parts = [p for p in [detalhe, favorecido] if p and p not in ("nan",)]
            descricao = " | ".join(parts) if parts else tipo_raw

        # ── CATEGORIA ─────────────────────────────────────────────────────────
        categoria_raw = str(
            row.get(self.COL_CATEGORIA) or
            row.get(self.COL_CATEGORIA_ALT) or ""
        ).strip()

        id_categoria = (
            self._categoria_map_local.get(categoria_raw) or
            CATEGORIA_MAP.get(categoria_raw)
        )

        if id_categoria is None:
            id_categoria = CATEGORIA_DEFAULT
            if categoria_raw and categoria_raw not in ("nan",):
                avisos.append(f"Categoria '{categoria_raw}' sem ID — usando padrão")
                status = max(status, self.STATUS_ATENCAO, key=self._status_weight)

        # ── FORMA DE PAGAMENTO ────────────────────────────────────────────────
        forma_raw = str(row.get(self.COL_FORMA_PGTO, "") or "").strip().upper()

        # Busca ID da forma de pagamento — prioridade: API > estatico
        modo_pgto_id = (
            self._payment_methods_local.get(forma_raw) or
            self._payment_methods_local.get(forma_raw.upper())
        )

        import logging as _log2
        _log2.getLogger("magical_conciliacao").info(
            f"[ERP] forma_raw='{forma_raw}' | pm_local={self._payment_methods_local} | id_encontrado={modo_pgto_id}")

        if modo_pgto_id:
            modo_pgto = int(modo_pgto_id)
        else:
            # Fallback estatico
            modo_pgto = FORMA_PGTO_MAP.get(forma_raw, FORMA_PGTO_DEFAULT)
            if forma_raw and forma_raw not in FORMA_PGTO_MAP:
                avisos.append(f"Forma '{forma_raw}' nao mapeada — usando padrao")
                status = max(status, self.STATUS_ATENCAO, key=self._status_weight)
            elif not modo_pgto_id:
                avisos.append(f"Forma '{forma_raw}' — usando mapa estatico (ID={modo_pgto})")
                status = max(status, self.STATUS_ATENCAO, key=self._status_weight)

        # ── PAYLOAD ───────────────────────────────────────────────────────────
        payload = {
            "datapagamento":   data_fmt,
            "valor":           round(valor, 2),
            "pago":            "nao",
            "tipocobranca":    tipocobranca,
            "descricao":       descricao[:200],
            "idcategoria":     id_categoria,
            "mododepagamento": modo_pgto,
        }
        if id_evento:
            payload["idevento"] = id_evento

        return {
            "_idx":               idx,
            "_status":            status,
            "_aviso":             " | ".join(avisos) if avisos else "",
            "_payload":           payload,
            "_data_fmt":          data_fmt,
            "_tipo_raw":          tipo_raw,
            "_id_evento_display": id_evento_display,
            "_categoria":         categoria_raw,
            "_id_categoria":      id_categoria,
            "_detalhe":           detalhe,
            "_valor":             valor,
            "_favorecido":        favorecido,
            "_forma_raw":         forma_raw,
            "_descricao":         descricao,
        }

    @staticmethod
    def _status_weight(s):
        return {ErpLancamentoWindow.STATUS_OK: 0,
                ErpLancamentoWindow.STATUS_ATENCAO: 1,
                ErpLancamentoWindow.STATUS_JA_LANCADO: 1,
                ErpLancamentoWindow.STATUS_BLOQUEIO: 2}.get(s, 0)

    # =========================================================================
    # POPUP DATAS PASSADAS
    # =========================================================================

    def _popup_selecionar_passadas(self, df_passadas: pd.DataFrame):
        """Popup com checkboxes para selecionar linhas passadas a incluir hoje."""
        from datetime import date as _date
        hoje_str = _date.today().strftime("%Y-%m-%d")
        hoje_fmt = _date.today().strftime("%d/%m/%Y")

        pop = tk.Toplevel(self.top)
        pop.title("Datas Retroativas")
        pop.geometry("680x420")
        pop.resizable(True, True)
        pop.grab_set()

        # Header
        hdr = tk.Frame(pop, bg="#92400e", pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text=f"⚠  {len(df_passadas)} linha(s) com data retroativa",
                 bg="#92400e", fg="white",
                 font=("Arial", 11, "bold")).pack()
        tk.Label(hdr,
                 text=f"Marque as que deseja lançar com a data de hoje ({hoje_fmt})",
                 bg="#92400e", fg="#fef3c7",
                 font=("Arial", 9)).pack()

        # Lista com checkboxes
        list_frame = tk.Frame(pop)
        list_frame.pack(fill="both", expand=True, padx=10, pady=8)

        # Cabecalho
        cab = tk.Frame(list_frame, bg="#f1f5f9")
        cab.pack(fill="x")
        tk.Label(cab, text="", width=3, bg="#f1f5f9").pack(side="left")
        for txt, w in [("Data orig.", 90), ("Categoria", 160),
                        ("Descrição", 220), ("Valor", 90)]:
            tk.Label(cab, text=txt, bg="#f1f5f9", fg="#475569",
                     font=("Arial", 8, "bold"), width=w//7,
                     anchor="w").pack(side="left", padx=2)

        # Scroll
        canvas = tk.Canvas(list_frame, highlightthickness=0)
        sb = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)

        inner = tk.Frame(canvas)
        canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        vars_check = []
        idxs       = []

        for _, row in df_passadas.iterrows():
            try:
                data_orig = pd.to_datetime(
                    row[self.COL_DATA], dayfirst=True).strftime("%d/%m/%Y")
            except Exception:
                data_orig = str(row.get(self.COL_DATA, ""))[:10]

            cat   = str(row.get(self.COL_CATEGORIA) or
                        row.get(self.COL_CATEGORIA_ALT) or "")[:22]
            desc  = str(row.get(self.COL_LANCAMENTO) or
                        row.get(self.COL_DETALHE) or "")[:30]
            try:
                val = abs(float(str(row.get(self.COL_VALOR,0))
                    .replace(",",".").replace("R$","").strip()))
                val_fmt = f"R$ {val:,.2f}".replace(",","X").replace(".",",").replace("X",".")
            except Exception:
                val_fmt = "—"

            f = tk.Frame(inner, pady=2)
            f.pack(fill="x", padx=4)

            var = tk.BooleanVar(value=False)
            vars_check.append(var)
            idxs.append(row.name)

            ttk.Checkbutton(f, variable=var).pack(side="left")
            tk.Label(f, text=data_orig, fg="#b45309",
                     font=("Arial", 9), width=10, anchor="w").pack(side="left")
            tk.Label(f, text=cat, font=("Arial", 9),
                     width=22, anchor="w").pack(side="left", padx=2)
            tk.Label(f, text=desc, font=("Arial", 9),
                     width=30, anchor="w").pack(side="left", padx=2)
            tk.Label(f, text=val_fmt, font=("Arial", 9, "bold"),
                     fg="#065f46", width=11, anchor="e").pack(side="left")

        # Selecionar todos
        sel_frame = tk.Frame(pop)
        sel_frame.pack(fill="x", padx=10)

        def _sel_todos():
            for v in vars_check:
                v.set(True)
        def _des_todos():
            for v in vars_check:
                v.set(False)

        ttk.Button(sel_frame, text="Marcar todas",
                   command=_sel_todos).pack(side="left", padx=4)
        ttk.Button(sel_frame, text="Desmarcar todas",
                   command=_des_todos).pack(side="left")

        # Rodape
        rod = tk.Frame(pop, pady=8)
        rod.pack(fill="x", padx=10)

        tk.Label(rod,
                 text=f"As selecionadas serão lançadas com data: {hoje_fmt}",
                 fg="#065f46", font=("Arial", 9, "italic")).pack(side="left")

        def _confirmar():
            selecionadas = [idxs[i] for i, v in enumerate(vars_check) if v.get()]
            if selecionadas:
                # Cria copia com data substituida por hoje
                df_sel = df_passadas.loc[selecionadas].copy()
                df_sel[self.COL_DATA] = pd.Timestamp.today().normalize()
                extra = self._validate_all(df_sel.reset_index(drop=True))
                self.df_preview = pd.concat(
                    [self.df_preview, extra], ignore_index=True)
                self._refresh_preview()
                # Atualiza totais
                ok    = (self.df_preview["_status"] == self.STATUS_OK).sum()
                atenc = (self.df_preview["_status"] == self.STATUS_ATENCAO).sum()
                bloq  = (self.df_preview["_status"] == self.STATUS_BLOQUEIO).sum()
                total_val = self.df_preview[
                    self.df_preview["_status"].isin(
                        [self.STATUS_OK, self.STATUS_ATENCAO])]["_valor"].sum()
                total_val_fmt = (f"R$ {total_val:,.2f}"
                    .replace(",","X").replace(".",",").replace("X","."))
                self.summary_label.config(
                    text=f"Total: {len(self.df_preview)}  |  ✓ {ok}  ⚠ {atenc}  ✗ {bloq}")
                self.total_valor_label.config(
                    text=f"Total a lançar: {total_val_fmt}")
                if ok + atenc > 0:
                    self.btn_lancar.config(state="normal")
            pop.destroy()

        ttk.Button(rod, text="Ignorar todas",
                   command=pop.destroy).pack(side="right", padx=4)
        ttk.Button(rod, text="Incluir selecionadas →",
                   command=_confirmar).pack(side="right")

    # =========================================================================
    # PREVIEW
    # =========================================================================

    def _refresh_preview(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        if self.df_preview is None or self.df_preview.empty:
            return

        filtro = self.filter_status_var.get()
        df = self.df_preview
        if filtro != "TODOS":
            df = df[df["_status"] == filtro]

        for _, row in df.iterrows():
            cat_display = row["_categoria"][:22]
            id_cat = row.get("_id_categoria", 1)
            cat_display += f" [{id_cat}]" if id_cat != 1 else " [?]"

            desc_display = row.get("_descricao", row["_detalhe"])[:45]

            self.tree.insert(
                "", "end",
                values=(
                    row["_status"],
                    row["_idx"] + 5,
                    row["_data_fmt"],
                    row["_id_evento_display"],
                    cat_display,
                    desc_display,
                    f"R$ {row['_valor']:,.2f}".replace(",","X").replace(".",",").replace("X","."),
                    row["_forma_raw"],
                    row["_aviso"][:90],
                ),
                tags=(row["_status"],)
            )

    # =========================================================================
    # LANÇAMENTO
    # =========================================================================

    def _confirm_and_launch(self):
        if self.df_preview is None or self.df_preview.empty:
            return

        dry      = self.dry_run_var.get()
        partner  = self.partner_var.get().strip()
        token    = self._current_token
        url_base = self._current_url.rstrip("/")

        if not partner:
            messagebox.showwarning("Aviso", "Selecione o parceiro antes de lançar.")
            return

        if not dry and not token:
            messagebox.showwarning("Aviso",
                f"Token da API não configurado para '{partner}'.\n\n"
                "Acesse Configurações → API MeEventos.")
            return

        if not dry and not url_base:
            messagebox.showwarning("Aviso",
                f"URL da API não configurada para '{partner}'.")
            return

        df_to_send = self.df_preview[
            self.df_preview["_status"].isin([self.STATUS_OK, self.STATUS_ATENCAO])
        ]
        ja_lancados = (self.df_preview["_status"] == self.STATUS_JA_LANCADO).sum()
        n    = len(df_to_send)
        bloq = (self.df_preview["_status"] == self.STATUS_BLOQUEIO).sum()

        modo_txt = "SIMULAÇÃO" if dry else "REAL"
        ja_lancados_count = (self.df_preview["_status"] == self.STATUS_JA_LANCADO).sum()
        msg = (
            f"Modo: {modo_txt}\n\n"
            f"Serão enviadas: {n} despesas (novas)\n"
            f"Já lançadas (ignoradas): {ja_lancados_count}\n"
            f"Bloqueadas (ignoradas): {bloq}\n\n"
            f"{'[SIMULAÇÃO — nenhum dado será enviado]' if dry else 'Os dados serão ENVIADOS para a API do MeEventos.'}\n\n"
            f"Confirma?"
        )
        if not messagebox.askyesno("Confirmar lançamento", msg):
            return

        self.btn_lancar.config(state="disabled")
        self.progress["value"] = 0
        self._set_status("Iniciando lançamento...")

        file_name = Path(self.file_path).name if self.file_path else "desconhecido"
        batch_id = self.repo.save_erp_launch_batch(
            partner_name  = partner,
            file_name     = file_name,
            file_path     = self.file_path or "",
            total_rows    = len(self.df_preview),
            total_enviado = 0,
            total_simulado= 0,
            total_erro    = 0,
            dry_run       = dry,
        )

        threading.Thread(
            target=self._launch_worker,
            args=(df_to_send, url_base, token, dry, partner, batch_id),
            daemon=True
        ).start()

    def _launch_worker(self, df_to_send, url_base, token, dry_run, partner_name, batch_id):
        from logger import log as _log
        total   = len(df_to_send)
        results = []
        file_name = Path(self.file_path).name if self.file_path else "desconhecido"

        for i, (_, row) in enumerate(df_to_send.iterrows()):
            payload = row["_payload"]
            idx     = row["_idx"]

            if dry_run:
                result = {
                    "idx":     idx,
                    "status":  "SIMULADO",
                    "message": f"Payload validado: valor={payload.get('valor')} cat={payload.get('idcategoria')}",
                }
            else:
                result = self._post_to_api(url_base, token, payload, idx)

                # Rate limit: se 429, aguarda e tenta novamente
                if result.get("status") == "ERRO_API" and "429" in str(result.get("message","")):
                    import time
                    self.top.after(0, lambda: self._set_status(
                        "Rate limit atingido — aguardando 10s...", "#b45309"))
                    time.sleep(10)
                    result = self._post_to_api(url_base, token, payload, idx)

                # Delay entre requisicoes para evitar rate limit (0.3s)
                import time
                time.sleep(0.6)

            _log.lancamento_item(
                parceiro  = partner_name,
                linha     = idx + 5,
                status    = result["status"],
                id_api    = result.get("id_api", ""),
                payload   = payload if result["status"] == "ERRO_API" else None,
                mensagem  = result.get("message", ""),
            )

            try:
                self.repo.save_erp_launch_item(
                    batch_id    = batch_id,
                    linha_excel = idx + 5,
                    partner_name= partner_name,
                    payload     = payload,
                    status      = result["status"],
                    id_api      = result.get("id_api", ""),
                    mensagem    = result.get("message", ""),
                    categoria   = row.get("_categoria", ""),
                )
            except Exception:
                pass

            results.append(result)
            progress = int((i + 1) / total * 100)
            self.top.after(0, lambda p=progress, r=result, ri=row: self._on_row_launched(p, r, ri))

        ok    = sum(1 for r in results if r["status"] == "LANCADO")
        erros = sum(1 for r in results if r["status"] == "ERRO_API")
        sim   = sum(1 for r in results if r["status"] == "SIMULADO")
        valor_total = sum(
            abs(float(row["_payload"].get("valor") or 0))
            for _, row in df_to_send.iterrows()
            if row["_payload"].get("valor")
        )
        _log.lancamento(
            parceiro    = partner_name,
            planilha    = file_name,
            caminho     = self.file_path or "",
            ok          = ok if not dry_run else sim,
            erros       = erros,
            simulado    = dry_run,
            valor_total = valor_total,
        )

        self.top.after(0, lambda: self._on_launch_done(results, dry_run, batch_id))

    def _post_to_api(self, url_base, token, payload, idx) -> dict:
        from logger import log as _log
        try:
            import requests as _requests
        except ImportError:
            return {"idx": idx, "status": "ERRO_API",
                    "message": "Pacote 'requests' não instalado."}

        try:
            resp = _requests.post(
                f"{url_base}/api/v1/financial",
                headers={
                    "Authorization": token,
                    "Content-Type":  "application/json",
                    "Accept":        "application/json",
                },
                json=payload,
                timeout=15,
            )
            if resp.status_code in (200, 201):
                data = resp.json()
                id_criado = ""
                if isinstance(data.get("data"), list) and data["data"]:
                    id_criado = data["data"][0].get("idmovimentacao", "")
                return {"idx": idx, "status": "LANCADO",
                        "message": f"ID criado: {id_criado}", "id_api": id_criado}
            else:
                return {"idx": idx, "status": "ERRO_API",
                        "message": f"HTTP {resp.status_code}: {resp.text[:120]}"}
        except Exception as exc:
            return {"idx": idx, "status": "ERRO_API", "message": str(exc)[:120]}

    def _on_row_launched(self, progress, result, row_series):
        self.progress["value"] = progress
        self._set_status(f"Lançando... {progress}%")

        for item in self.tree.get_children():
            vals = self.tree.item(item, "values")
            linha_excel = row_series["_idx"] + 5
            if vals and str(vals[1]) == str(linha_excel):
                st = result.get("status", "")
                msg = result.get("message", "")
                novo_status = ("✓ Lançado" if st == "LANCADO"
                               else "✓ Simulado" if st == "SIMULADO"
                               else "✗ Erro API")
                tag = "LANCADO" if st in ("LANCADO", "SIMULADO") else "ERRO_API"
                new_vals = list(vals)
                new_vals[0] = novo_status
                new_vals[8] = msg[:80] if msg else ""
                self.tree.item(item, values=new_vals, tags=(tag,))
                break

    def _on_launch_done(self, results, dry_run, batch_id=None):
        self.progress["value"] = 100
        lancados = sum(1 for r in results if r.get("status") in ("LANCADO", "SIMULADO"))
        erros    = sum(1 for r in results if r.get("status") == "ERRO_API")

        modo = "simuladas" if dry_run else "lançadas"
        self._set_status(
            f"Concluído — {lancados} {modo}, {erros} com erro.",
            "#166534" if erros == 0 else "#b91c1c"
        )
        self.btn_lancar.config(state="normal")

        try:
            modo_log = "SIMULAÇÃO" if dry_run else "REAL"
            self.repo.log("INFO", f"Lançamento ERP {modo_log} concluído",
                          f"batch_id={batch_id} | {lancados} {modo} | {erros} erro(s)")
        except Exception:
            pass

        self._notificar_lancamento(lancados, erros, batch_id, dry_run=dry_run)
        self._atualizar_pendencia_pos_lancamento(lancados, erros, dry_run)

        msg = (
            f"Processo concluído!\n\n"
            f"{'Simuladas' if dry_run else 'Lançadas'}: {lancados}\n"
            f"Erros: {erros}\n"
        )
        if dry_run:
            msg += "\nNenhum dado foi enviado para a API (modo simulação)."

        messagebox.showinfo("Resultado", msg)

    def _atualizar_pendencia_pos_lancamento(self, lancados, erros, dry_run):
        planilha_id = getattr(self, "_pendencia_id", None)
        casa        = getattr(self, "_pendencia_casa", "")
        if not planilha_id:
            return

        def _do():
            try:
                import sqlite3
                db_path = str(self.repo.db.db_path)
                def get_cfg(chave, default=""):
                    try:
                        with sqlite3.connect(db_path) as conn:
                            conn.execute("CREATE TABLE IF NOT EXISTS nuvem_config (chave TEXT PRIMARY KEY, valor TEXT)")
                            row = conn.execute("SELECT valor FROM nuvem_config WHERE chave=? LIMIT 1", (chave,)).fetchone()
                            return row[0] if row and row[0] else default
                    except Exception:
                        return default

                pb_url   = get_cfg("pb_url")
                pb_email = get_cfg("pb_email")
                pb_senha = get_cfg("pb_senha")
                meu_nome = get_cfg("meu_nome", "Sistema")

                if not pb_url:
                    return

                from cloud_sync import CloudSync
                cs = CloudSync(pb_url, pb_email, pb_senha)

                valor_total = 0.0
                try:
                    if self.df_preview is not None:
                        mask = self.df_preview["_status"] != self.STATUS_BLOQUEIO
                        valor_total = float(self.df_preview.loc[mask, "_valor"].sum())
                except Exception:
                    pass

                cs.confirmar_lancamento(
                    planilha_id = planilha_id,
                    casa        = casa,
                    usuario     = meu_nome,
                    ok          = lancados,
                    erros       = erros,
                    valor_total = valor_total,
                )
            except Exception as e:
                import logging
                logging.getLogger("magical_conciliacao").error(
                    f"[PENDENCIA] Erro ao atualizar status: {e}")

        threading.Thread(target=_do, daemon=True).start()

    def _notificar_lancamento(self, lancados, erros, batch_id, dry_run=False):
        def _enviar():
            try:
                import sqlite3
                db_path = str(self.repo.db.db_path)
                def get_cfg(chave, default=""):
                    try:
                        with sqlite3.connect(db_path) as conn:
                            conn.execute("CREATE TABLE IF NOT EXISTS nuvem_config (chave TEXT PRIMARY KEY, valor TEXT)")
                            row = conn.execute("SELECT valor FROM nuvem_config WHERE chave=? LIMIT 1", (chave,)).fetchone()
                            return row[0] if row and row[0] else default
                    except Exception:
                        return default

                evo_url       = get_cfg("evo_url")
                evo_key       = get_cfg("evo_key")
                instancia     = get_cfg("evo_instancia", "wanderley")
                num_marcielo  = get_cfg("num_marcielo")
                num_wanderley = get_cfg("num_wanderley")
                meu_nome      = get_cfg("meu_nome", "Sistema")

                if not evo_url or not evo_key:
                    return

                from notificador import Notificador
                n = Notificador(evo_url, evo_key, instancia)
                if not n.status_instancia().get("conectado"):
                    return

                parceiro  = self.partner_var.get() if hasattr(self, 'partner_var') else ""
                file_name = Path(self.file_path).name if self.file_path else "planilha"

                valor_total = 0.0
                try:
                    if self.df_preview is not None:
                        mask = self.df_preview["_status"] != self.STATUS_BLOQUEIO
                        valor_total = float(self.df_preview.loc[mask, "_valor"].sum())
                except Exception:
                    pass

                prefixo = "[SIMULACAO] " if dry_run else ""

                if erros == 0:
                    msg = (
                        f"*{prefixo}Magical — Despesas Lancadas no MeEventos*\n\n"
                        f"Casa: {parceiro}\n"
                        f"Planilha: {file_name}\n"
                        f"Lancado por: {meu_nome}\n"
                        f"Total: R$ {valor_total:,.2f}\n"
                        f"Lancamentos: {lancados}\n\n"
                        f"CNAB disponivel para geracao."
                    ).replace(",","X").replace(".",",").replace("X",".")
                else:
                    msg = (
                        f"*{prefixo}Magical — ERP com Erros*\n\n"
                        f"Casa: {parceiro}\n"
                        f"Lancados: {lancados} | Erros: {erros}\n\n"
                        f"Verifique os erros no sistema."
                    )

                for num in [num_marcielo, num_wanderley]:
                    if num:
                        n.enviar_contato(num, msg)

            except Exception:
                pass

        threading.Thread(target=_enviar, daemon=True).start()